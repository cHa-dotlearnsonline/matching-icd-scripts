import openpyxl 
from openpyxl import Workbook

#load all the workbooks I am using today:
HIA_data_wb = openpyxl.load_workbook("matched_diagnosis_hia1b_final.xlsx")
ICD11_Oid_data_wb = openpyxl.load_workbook("CAREPRO-ICD11.xlsx")

HIA_data_ws = HIA_data_wb["Sheet1"]

ICD11_Oid_data_ws = ICD11_Oid_data_wb["Sheet1"]

# let's load the data from HIA into a dictionary that matches the name to the code:
all_hia_rows = HIA_data_ws.max_row 
A = "A"
B = "B"

all_oid_rows = ICD11_Oid_data_ws.max_row
matched_material2 = {}
for i in range(2, all_oid_rows):
    matched_material2[" ".join(str(ICD11_Oid_data_ws[f"C{i}"].value).lower().split())] = str(ICD11_Oid_data_ws[f"A{i}"].value).lower()

new_workbook = Workbook()
new_worksheet = new_workbook.create_sheet(title="Matched to Oid")

new_worksheet = new_workbook["Matched to Oid"]
total_hia_rows = HIA_data_ws.max_row
for i in range(1, total_hia_rows):  
    matcher = HIA_data_ws[f"C{i}"].value
    new_worksheet.cell(row=i, column=1).value = HIA_data_ws[f"A{i}"].value
    if matcher is not None:
        matches = HIA_data_ws[f"C{i}"].value.split("+")
        matches1 = [" ".join(diagnosis.lower().split()) for diagnosis in matches]
        my_diagnosis = []
        for match in matches1:
            if match in matched_material2.keys():
                my_diagnosis.append(matched_material2[match])
        new_worksheet.cell(row=i, column=2).value = ",".join(my_diagnosis)

new_workbook.save("HIA_1_B_Matched_to_Oid(5)_final.xlsx")
HIA_data_wb.close()
ICD11_Oid_data_wb.close()
