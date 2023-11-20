from fuzzywuzzy import fuzz 
from fuzzywuzzy import process
import openpyxl 
from openpyxl import Workbook

HIA_data_wb = openpyxl.load_workbook("HIA 1 B diagnosis.xlsx")
ICD11_data_wb = openpyxl.load_workbook("ICD SC evo.xlsx")

HIA_data_ws = HIA_data_wb["Sheet1"]

ICD11_data_ws = ICD11_data_wb["icd11A"]

# let's load all the data from the excel sheet for HIA into a list
all_rows = HIA_data_ws.max_row
column = "A"
all_HIA_diagnosis = []
for i in range(2,all_rows):
    all_HIA_diagnosis.append(" ".join(str(HIA_data_ws[f"{column}{i}"].value).lower().split()))

# let's load all the diagnosis for ICD11 into a list as well
all_icd11_rows = ICD11_data_ws.max_row
all_ICD11_diagnosis = []
for i in range(1, all_icd11_rows):
    all_ICD11_diagnosis.append(" ".join(str(ICD11_data_ws[f"B{i}"].value).lower().split()))
    
new_workbook = Workbook()
tracker = 1
for diagnosis in all_HIA_diagnosis:
    worksheet1 = new_workbook["Sheet"]
    my_boy = process.extract(diagnosis, all_ICD11_diagnosis, limit=50)
    my_rexo =[]
    for boy in my_boy:
        if int(boy[1]) >= 88:
            my_rexo.append(boy[0])
    my_rexo = [ str(boy) for boy in my_rexo]
    worksheet1.append([diagnosis])
    worksheet1[f"B{tracker}"].value = " + ".join(my_rexo)
    tracker += 1
new_workbook.save("matched_diagnosis_hia_1_b(3).xlsx")

HIA_data_wb.close()
ICD11_data_wb.close()
